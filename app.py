from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
import openpyxl
import pandas as pd
import json
from datetime import datetime

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')
app.config['DATABASE'] = 'checklist.db'

csrf = CSRFProtect(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, id, username, name, email, department, role):
        self.id = id
        self.username = username
        self.name = name
        self.email = email
        self.department = department
        self.role = role

def get_db():
    db = sqlite3.connect(app.config['DATABASE'])
    db.row_factory = sqlite3.Row
    return db

def init_db():
    db = get_db()
    
    db.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        email TEXT NOT NULL,
        department TEXT NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL
    )''')
    
    db.execute('''CREATE TABLE IF NOT EXISTS checklist_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sheet_name TEXT NOT NULL,
        row_index INTEGER NOT NULL,
        col_index INTEGER NOT NULL,
        value TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    db.execute('''CREATE TABLE IF NOT EXISTS checklist_structure (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sheet_name TEXT NOT NULL,
        headers TEXT NOT NULL,
        total_rows INTEGER NOT NULL,
        total_cols INTEGER NOT NULL,
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    db.execute('''CREATE TABLE IF NOT EXISTS worksheets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sheet_name TEXT UNIQUE NOT NULL,
        display_order INTEGER NOT NULL,
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    admin_exists = db.execute("SELECT * FROM users WHERE username = 'admin'").fetchone()
    if not admin_exists:
        admin_password = generate_password_hash('admin123')
        db.execute("INSERT INTO users (username, name, email, department, password, role) VALUES (?, ?, ?, ?, ?, ?)",
                   ('admin', 'Administrator', 'admin@example.com', 'Admin', admin_password, 'admin'))
    
    user_exists = db.execute("SELECT * FROM users WHERE username = 'user'").fetchone()
    if not user_exists:
        user_password = generate_password_hash('user123')
        db.execute("INSERT INTO users (username, name, email, department, password, role) VALUES (?, ?, ?, ?, ?, ?)",
                   ('user', 'Regular User', 'user@example.com', 'Operations', user_password, 'user'))
    
    db.commit()
    db.close()

@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    user_data = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    db.close()
    
    if user_data:
        return User(user_data['id'], user_data['username'], user_data['name'], 
                   user_data['email'], user_data['department'], user_data['role'])
    return None

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('user_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
@csrf.exempt
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        db = get_db()
        user_data = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        db.close()
        
        if user_data and check_password_hash(user_data['password'], password):
            user = User(user_data['id'], user_data['username'], user_data['name'],
                       user_data['email'], user_data['department'], user_data['role'])
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('user_dashboard'))
    
    db = get_db()
    users = db.execute("SELECT * FROM users").fetchall()
    db.close()
    
    return render_template('admin_dashboard.html', users=users)

@app.route('/admin/user/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    username = request.form.get('username')
    name = request.form.get('name')
    email = request.form.get('email')
    department = request.form.get('department')
    password = request.form.get('password')
    role = request.form.get('role')
    
    db = get_db()
    try:
        hashed_password = generate_password_hash(password)
        db.execute("INSERT INTO users (username, name, email, department, password, role) VALUES (?, ?, ?, ?, ?, ?)",
                   (username, name, email, department, hashed_password, role))
        db.commit()
        flash('User added successfully', 'success')
    except sqlite3.IntegrityError:
        flash('Username already exists', 'error')
    finally:
        db.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/user/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    if user_id == current_user.id:
        flash('Cannot delete your own account', 'error')
        return redirect(url_for('admin_dashboard'))
    
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    db.close()
    
    flash('User deleted successfully', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/user/dashboard')
@login_required
def user_dashboard():
    return render_template('user_dashboard.html')

@app.route('/checklist')
@app.route('/checklist/<sheet_name>')
@login_required
def checklist(sheet_name=None):
    db = get_db()
    
    worksheets = db.execute("SELECT * FROM worksheets ORDER BY display_order").fetchall()
    
    if not worksheets:
        db.close()
        return render_template('checklist.html', headers=[], data=[], worksheets=[], current_sheet=None)
    
    if not sheet_name:
        sheet_name = worksheets[0]['sheet_name']
    
    structure = db.execute("SELECT * FROM checklist_structure WHERE sheet_name = ?", (sheet_name,)).fetchone()
    
    if structure:
        headers = json.loads(structure['headers'])
        total_rows = structure['total_rows']
        total_cols = structure['total_cols']
        
        data_rows = db.execute("SELECT * FROM checklist_data WHERE sheet_name = ?", (sheet_name,)).fetchall()
        
        data_grid = [['' for _ in range(total_cols)] for _ in range(total_rows)]
        
        for row in data_rows:
            if row['row_index'] < total_rows and row['col_index'] < total_cols:
                data_grid[row['row_index']][row['col_index']] = row['value'] or ''
        
        db.close()
        return render_template('checklist.html', headers=headers, data=data_grid, worksheets=worksheets, current_sheet=sheet_name)
    else:
        db.close()
        return render_template('checklist.html', headers=[], data=[], worksheets=worksheets, current_sheet=sheet_name)

@app.route('/upload_excel', methods=['POST'])
@login_required
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    try:
        wb = openpyxl.load_workbook(file)
        
        db = get_db()
        
        db.execute("DELETE FROM checklist_data")
        db.execute("DELETE FROM checklist_structure")
        db.execute("DELETE FROM worksheets")
        
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else '')
            
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append([cell if cell is not None else '' for cell in row])
            
            total_rows = len(data)
            total_cols = len(headers)
            
            db.execute("INSERT INTO worksheets (sheet_name, display_order) VALUES (?, ?)",
                       (sheet_name, idx))
            
            db.execute("INSERT INTO checklist_structure (sheet_name, headers, total_rows, total_cols) VALUES (?, ?, ?, ?)",
                       (sheet_name, json.dumps(headers), total_rows, total_cols))
            
            for row_idx, row in enumerate(data):
                for col_idx, value in enumerate(row):
                    db.execute("INSERT INTO checklist_data (sheet_name, row_index, col_index, value) VALUES (?, ?, ?, ?)",
                               (sheet_name, row_idx, col_idx, str(value)))
        
        db.commit()
        db.close()
        
        return jsonify({'success': True, 'message': f'Uploaded {len(wb.sheetnames)} worksheets successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/save_cell', methods=['POST'])
@login_required
def save_cell():
    data = request.json
    sheet_name = data.get('sheet_name')
    row = data.get('row')
    col = data.get('col')
    value = data.get('value')
    
    db = get_db()
    
    existing = db.execute("SELECT * FROM checklist_data WHERE sheet_name = ? AND row_index = ? AND col_index = ?", 
                         (sheet_name, row, col)).fetchone()
    
    if existing:
        db.execute("UPDATE checklist_data SET value = ?, updated_at = ? WHERE sheet_name = ? AND row_index = ? AND col_index = ?",
                   (value, datetime.now(), sheet_name, row, col))
    else:
        db.execute("INSERT INTO checklist_data (sheet_name, row_index, col_index, value) VALUES (?, ?, ?, ?)",
                   (sheet_name, row, col, value))
    
    db.commit()
    db.close()
    
    return jsonify({'success': True})

@app.route('/download_excel')
@login_required
def download_excel():
    try:
        db = get_db()
        
        template_file = 'attached_assets/CR Check List - Latest Format_1760583521780.xlsx'
        
        if os.path.exists(template_file):
            wb = openpyxl.load_workbook(template_file)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
        
        worksheets = db.execute("SELECT * FROM worksheets ORDER BY display_order").fetchall()
        
        for worksheet in worksheets:
            sheet_name = worksheet['sheet_name']
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            structure = db.execute("SELECT * FROM checklist_structure WHERE sheet_name = ?", (sheet_name,)).fetchone()
            
            if structure:
                headers = json.loads(structure['headers'])
                total_rows = structure['total_rows']
                total_cols = structure['total_cols']
                
                data_rows = db.execute("SELECT * FROM checklist_data WHERE sheet_name = ?", (sheet_name,)).fetchall()
                
                data_grid = [['' for _ in range(total_cols)] for _ in range(total_rows)]
                
                for row in data_rows:
                    if row['row_index'] < total_rows and row['col_index'] < total_cols:
                        data_grid[row['row_index']][row['col_index']] = row['value'] or ''
                
                for row_idx, row_data in enumerate(data_grid):
                    for col_idx, cell_value in enumerate(row_data):
                        ws.cell(row=row_idx+2, column=col_idx+1, value=cell_value)
        
        db.close()
        
        excel_filename = f'attached_assets/CR_Checklist_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(excel_filename)
        
        return send_file(excel_filename, 
                        as_attachment=True, 
                        download_name=f'CR_Checklist_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
