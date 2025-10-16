import sqlite3
import openpyxl
import json

excel_file = 'attached_assets/CR Check List - Latest Format_1760583521780.xlsx'
db_file = 'checklist.db'

try:
    wb = openpyxl.load_workbook(excel_file)
    
    db = sqlite3.connect(db_file)
    
    db.execute("DELETE FROM checklist_data")
    db.execute("DELETE FROM checklist_structure")
    db.execute("DELETE FROM worksheets")
    
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else '')
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append([cell if cell is not None else '' for cell in row])
        
        total_rows = len(data)
        total_cols = len(headers)
        
        db.execute("INSERT INTO worksheets (sheet_name, display_order) VALUES (?, ?)",
                   (sheet_name, idx))
        
        db.execute("INSERT INTO checklist_structure (sheet_name, headers, total_rows, total_cols) VALUES (?, ?, ?, ?)",
                   (sheet_name, json.dumps(headers), total_rows, total_cols))
        
        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                db.execute("INSERT INTO checklist_data (sheet_name, row_index, col_index, value) VALUES (?, ?, ?, ?)",
                           (sheet_name, row_idx, col_idx, str(value)))
        
        print(f"Loaded worksheet: {sheet_name} ({total_rows} rows, {total_cols} cols)")
    
    db.commit()
    db.close()
    
    print(f"\nSuccessfully loaded {len(wb.sheetnames)} worksheets into database")
except Exception as e:
    print(f"Error loading Excel file: {e}")
